# Created by Omer Atias
# TopStocks - The app takes the five stocks that had the most significant growth in the last trading day from the website and automatically sends them to your email!

import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

# Email configuration for Gmail
sender_email = 'sender@mail.com'  # Replace with your Gmail address
receiver_email = 'recipient@mail.com'  # Replace with recipient's email
password = 'password'  # Replace with your App Password or Gmail password
smtp_server = 'smtp.gmail.com'
smtp_port = 587

# URL of the Yahoo Finance page for stock gainers
url = 'https://finance.yahoo.com/markets/stocks/gainers/'

# Make a GET request to fetch the raw HTML content
response = requests.get(url)

# Check if the request was successful
if response.status_code == 200:
    # Parse the HTML content using BeautifulSoup
    soup = BeautifulSoup(response.content, 'html.parser')

    # Find the table containing the stock gainers
    table = soup.find('table')

    # Get the rows of the table
    rows = table.find_all('tr')[1:6]  # Skip the header and take the first 5 rows

    # List to hold the stock data
    stock_data = []

    # Loop through each row to extract stock details
    for row in rows:
        cols = row.find_all('td')
        stock_info = {
            'symbol': cols[0].text.strip(),
            'name': cols[1].text.strip(),
            'price': cols[2].text.strip(),
            'change': cols[3].text.strip(),
            'percent_change': cols[4].text.strip()
        }
        stock_data.append(stock_info)  # Append the stock_info dictionary to the list

    # Create a DataFrame from the list of stock data
    df = pd.DataFrame(stock_data)

    # Define the path for the Excel file (Desktop path)
    excel_path = os.path.join(os.path.expanduser('~'), 'Desktop', 'top_gaining_stocks.xlsx')

    # Save the DataFrame to an Excel file
    df.to_excel(excel_path, index=False, sheet_name='Top Gainers')

    # Open the Excel file to modify it
    workbook = load_workbook(excel_path)  # Load the workbook
    sheet = workbook.active

    # Set the header row to be bold and filled with blue
    fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)  # White font color

    for cell in sheet[1]:  # Iterate over the header row
        cell.fill = fill
        cell.font = font

    # Save the modified Excel file
    workbook.save(excel_path)

    # Prepare email
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = 'Top 5 Stocks Today'

    # Email body
    body = "Hello,\n\nHere is a list of the Top 5 stocks that were most increased in today's day-trade according to Yahoo! Finance.\n\nTopStocks App."
    msg.attach(MIMEText(body, 'plain'))

    # Attach the Excel file
    with open(excel_path, 'rb') as attachment:
        part = MIMEApplication(attachment.read(), Name=os.path.basename(excel_path))
        part['Content-Disposition'] = f'attachment; filename="{os.path.basename(excel_path)}"'
        msg.attach(part)

    # Send the email
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()  # Upgrade the connection to a secure encrypted SSL/TLS connection
        server.login(sender_email, password)  # Login to your email account
        server.send_message(msg)  # Send the email

    print(f'Top 5 gaining stocks saved to {excel_path} and email sent to {receiver_email}')
else:
    print(f'Failed to retrieve data. Status code: {response.status_code}')
